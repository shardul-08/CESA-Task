from typing import List
from openpyxl import load_workbook

def load_emails_from_excel(path: str) -> List[str]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    emails = []
    for row in ws.iter_rows(min_row=1, max_col=1):
        cell = row[0].value
        if cell and isinstance(cell, str) and "@" in cell:
            emails.append(cell.strip())
    return emails
